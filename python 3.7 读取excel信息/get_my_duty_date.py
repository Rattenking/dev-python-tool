#!/usr/bin/env python
"""
@Author  ：Rattenking
@Date    ：2021/06/02 10:19
@CSDN	 ：https://blog.csdn.net/m0_38082783
"""

import openpyxl
import time

def get_my_duty_date():
  dutys = []
  book = openpyxl.load_workbook('duty.xlsx',data_only=True)
  sheet = book.active
  all_data = book.get_sheet_by_name("日常加班")
  none_str = ''.join([str(None).ljust(20) for c in range(1,all_data.max_column+1)])
  for r in range(1,all_data.max_row + 1):
    cur_str = ''.join([str(all_data.cell(row=r,column=c).value).ljust(20) for c in range(1,all_data.max_column+1)])
    if cur_str.find("梁亚军") >= 0:
      dutys.append({
        "date": all_data.cell(row=r,column=2).value,
        "name": all_data.cell(row=r,column=3).value
      })
    elif cur_str.find(none_str) >= 0:
      break
  return dutys

def create_my_duty_list(dutys):
  book = openpyxl.Workbook()
  sheet = book.active
  for i in range(len(dutys)):
    sheet.cell(row=1 + i, column=1).value = dutys[i].get("name")
    sheet.cell(row=1 + i, column=2).value = f'{dutys[i].get("date")}'
  book.save('my_duty.xlsx')

if __name__ == "__main__":
  start_time = int(round(time.time() * 1000))
  dutys = get_my_duty_date()
  create_my_duty_list(dutys)
  end_time = int(round(time.time() * 1000))
  print(f'本次提取值班表时间：{end_time - start_time}ms')